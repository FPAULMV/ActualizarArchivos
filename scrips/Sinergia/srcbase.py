from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import uuid
from urllib.parse import quote_plus 
import sqlalchemy
import pyodbc
import pandas as pd
import logging
from decouple import config
import os
import subprocess


import datetime as dt


# Configuración de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Variables de conexión (tomadas de variables de entorno)
server = config('server')
database = config('database')
user_name = config('user_name')
password = config('password')
driver = config('driver')

# Función para ejecutar consulta SQL
def solicitar(query: str) -> pd.DataFrame:
    """Crea una conexión a SQL Server donde ejecutará una consulta de tipo SELECT 
       y devuelve un dataframe con la información de la consulta."""
    
    conn_str = f"mssql+pymssql://{user_name}:{quote_plus(password)}@{server}/{database}"
    engine = sqlalchemy.create_engine(conn_str)
    try:
        resultado = pd.read_sql(query, engine)
        return resultado
    except Exception as e:
        logging.error(f"Error en la conexión (Revisar conexion, VPN o internet) o en la ejecución del query: {e}")
    finally:
        engine.dispose()


# Enviar registros de actividad a una tabla de SQL
def logs_update(start: str, end: str, file: str):
    """Cadena de texto para SQL Server que completa el query INSERT INTO --> VALUES (values)"""
    path = config('path_update')
    query_insert = f"""INSERT INTO Sinergia_Aux.dbo.log_actualizacion_archivos 
            ([Start], [End], [File], [Path]) 
            VALUES('{start}','{end}','{file}','{path}')"""
    conn_str = (
        f"DRIVER={driver};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"UID={user_name};"
        f"PWD={password}"
    )
    try:
        with pyodbc.connect(conn_str) as conn:  
            with conn.cursor() as cursor:
                cursor.execute(query_insert)
                cursor.commit()
    except pyodbc.Error as e:
        print("Ocurrió un error:", e)
        logging.error(f"Error en la conexión o en la ejecución del query: {e}")
    except Exception as e:
        logging.error(f"Ocurrió un error inesperado: {e}")


# Función para crear archivo Excel con múltiples hojas y consultas
def create_excel(queries_dict: dict, filename: str):
    path = config('path_update')
    outputpath = f"{path}\\{filename}.xlsx"
    tablestyle = "TableStyleMedium11"
    
    logging.info("Iniciando creación de archivo Excel")

    # Crear nuevo libro de trabajo Excel
    wb = Workbook()

    # Eliminar la hoja por defecto "Sheet" si existe
    default_sheet = wb['Sheet']
    if default_sheet:
        wb.remove(default_sheet)

    for sheetname, query in queries_dict.items():
        logging.info(f"Ejecutando consulta para hoja: {sheetname}")
        
        # Ejecutar la consulta SQL y obtener el resultado como DataFrame
        df = solicitar(query)
        if df.empty:
            logging.warning(f"La consulta para hoja '{sheetname}' no devolvió resultados, se omitirá.")
            continue
        
        # Convertir UUIDs a cadenas de texto
        df = df.map(lambda x: str(x).upper() if isinstance(x, uuid.UUID) else x)

        # Crear nueva hoja en el libro de trabajo
        ws = wb.create_sheet(title=sheetname)
        
        # Agregar datos al worksheet utilizando dataframe_to_rows
        for row in dataframe_to_rows(df, index=False, header=True):
            try:
                ws.append(row)
            except ValueError as e:
                logging.error(f"Error al agregar fila a la hoja '{sheetname}': {e}")

        # Formatear columnas de fechas
        date_columns = df.select_dtypes(include=['datetime64']).columns
        for col in date_columns:
            col_idx = df.columns.get_loc(col) + 1  # +1 porque Excel es 1-indexed
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                if cell.row != 1:  # Saltar el encabezado
                    cell.number_format = 'DD/MM/YYYY'
        
        # Crear tabla en la hoja
        logging.info(f"Creando tabla para la hoja: {sheetname}")
        tab = Table(displayName=sheetname, ref=ws.dimensions)
        logging.info("Aplicando estilos")
        style = TableStyleInfo(name=tablestyle, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    
    # Guardar el archivo Excel
    wb.save(outputpath)
    logging.info(f"Archivo Excel guardado")


def src_exect(queries,filename):
    start = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    print(f"""Cargando: ----- {filename} -----""")
    create_excel(queries, filename)
    end = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    logs_update(start, end, filename)
    logging.info("Proceso finalizado")


def activar_entorno_virtual():
    """
    Activa el entorno virtual en la ruta especificada.

    Args:
        ruta_entorno (str): Ruta al archivo `activate.bat` del entorno virtual.
    """
    ruta_entorno = config('VENV_PATH')

    if not os.path.exists(ruta_entorno):
        print(f"El archivo {ruta_entorno} no existe. Verifica la ruta.")
        return

    try:
        # Activar el entorno virtual
        subprocess.run(f'cmd /k "{ruta_entorno}"', shell=True)
        print("Entorno virtual activado. Puedes trabajar en él ahora.")
    except Exception as e:
        print(f"Error al activar el entorno virtual: {e}")


"""El siguiente es un ejemplo de como se debe de ejecutar el codigo, disponible para copiar y pegar en el nuevo proyecto:

# ----------------------------------------------------------------      
from srcbase import src_exect
import os

if __name__ == "__main__":
    # Parametros de entrada.
    # Agregar del lado izquierdo el nombre de la hoja y del lado derecho la consulta.
    queries = {
        "": ""
    }
    filename = os.path.basename(__file__).split('.')[0]

    # Ejecución
    src_exect(queries,filename)
# ----------------------------------------------------------------

"""
